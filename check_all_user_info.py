import os
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# ====================== 配置区（修改这里）======================
ROOT_DIR = r"D:\人员资质、身份证"  # 你的根目录
ROOT_DIR = r"C:\Users\101202304023\Desktop\工作\投标项目\用户身份证、毕业证、合同等信息"
ROOT_DIR = r"D:\用户身份证、毕业证、合同等信息"
SKIP_FOLDERS = {"新增资料", "所有照片"}  # 跳过的文件夹
EXCEL_SAVE_PATH = r"./人员资质汇总校验表.xlsx"  # 生成excel保存路径
# 表格输出字段顺序（表头名: json键名）
COLUMN_CONFIG = [
    ("姓名", "name"),
    ("身份证", "idcard"),
    ("合同", "contract"),
    ("毕业证", "graduationCertificate"),
    ("学位证", "degreeCertificate"),
    ("学信网备案/查询", "xuexinwang"),
    ("CISP证书", "cisp"),
    ("CISAW证书", "cisaw"),
    ("PMP", "pmp"),
    ("软考类证书", "ruankao"),
    ("未分类文件", "other"),

]
# ==============================================================

def to_list(val):
    """统一转为列表，单个字符串也包装成列表"""
    if val is None:
        return []
    if isinstance(val, str):
        return [val]
    if isinstance(val, list):
        return val
    return []

def check_person_folder(person_dir: Path):
    """
    单人目录校验
    返回：(人员名称, json数据, 错误信息列表)
    """
    errors = []
    json_file = person_dir / "info.json"
    person_name = person_dir.name

    # 1. 判断info.json是否存在
    if not json_file.exists():
        errors.append(f"【缺失文件】{person_name} 目录下无 info.json")
        return person_name, {}, errors

    # 2. 读取并解析json
    try:
        with open(json_file, "r", encoding="utf-8") as f:
            info_data = json.load(f)
    except Exception as e:
        errors.append(f"【JSON解析失败】{person_name} info.json 格式错误：{str(e)}")
        return person_name, {}, errors

    # 3. 收集json中记录的全部文件名
    json_all_files = set()
    for key in info_data:
        if key == "name":
            continue
        file_list = to_list(info_data[key])
        for fn in file_list:
            json_all_files.add(fn)

    # 4. 收集本地真实文件（排除info.json、子文件夹）
    local_all_files = set()
    for item in person_dir.iterdir():
        if item.is_dir() or item.name == "info.json":
            continue
        local_all_files.add(item.name)

    # 5. 校验：json有，但本地无文件
    missing_in_local = json_all_files - local_all_files
    for fn in missing_in_local:
        errors.append(f"【文件丢失】{person_name}：json记录[{fn}] 本地不存在")

    # 6. 校验：本地有，但json未收录
    extra_local = local_all_files - json_all_files
    for fn in extra_local:
        errors.append(f"【未录入JSON】{person_name}：本地文件[{fn}] 未写入info.json")

    return person_name, info_data, errors
import sys

def main():
    root_path = ROOT_DIR
    output_excel = EXCEL_SAVE_PATH

    # 访问特定的命令行参数
    if len(sys.argv) > 2:
        root_path = sys.argv[1]
        output_excel = sys.argv[2]
    else:
        print("Usage: python scan_all_user_info.py root_path excel_output")
        return

    root = Path(root_path)
    if not root.exists():
        print(f"错误：根目录不存在，请检查路径 -> {root_path}")
        return

    print(f"Root :'{root_path}', excel output:'{output_excel}'")


    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "人员资质清单"

    # 写入表头
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    headers = [col[0] for col in COLUMN_CONFIG]
    ws.append(headers)
    for col_idx in range(1, len(headers)+1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    total_person = 0
    has_error_count = 0

    # 遍历根目录下所有一级文件夹（人员文件夹）
    for entry in root.iterdir():
        if not entry.is_dir():
            continue
        dir_name = entry.name
        if dir_name in SKIP_FOLDERS:
            print(f"跳过目录：{dir_name}")
            continue

        total_person += 1
        name, data, err_list = check_person_folder(entry)

        # 打印错误信息
        if err_list:
            has_error_count += 1
            print("="*60)
            for err in err_list:
                print(err)
            print("="*60 + "\n")

        # 组装单行数据写入Excel
        row_data = []
        for header_name, json_key in COLUMN_CONFIG:
            val = data.get(json_key, "")
            # 列表转逗号分隔字符串
            if isinstance(val, list):
                val = "，".join(val)
            row_data.append(val)
        ws.append(row_data)

    # 自动调整列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if length > max_len:
                    max_len = length
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    # 保存Excel
    wb.save(output_excel)
    print(f"\n==================== 校验完成 ====================")
    print(f"总人员目录数：{total_person}")
    print(f"存在异常的目录数：{has_error_count}")
    print(f"汇总表格已生成：{output_excel}")

if __name__ == "__main__":
    main()